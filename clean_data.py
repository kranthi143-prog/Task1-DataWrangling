"""
Task 1: Data Immersion & Wrangling
ApexPlanet Data Analytics Internship

This script:
  1. Loads the raw sales dataset
  2. Profiles it to identify data quality issues (missing values, duplicates,
     inconsistent formatting, outliers)
  3. Cleans and transforms the data (imputation, ID correction, date
     standardization, feature engineering)
  4. Outputs a final, analysis-ready dataset

Run:  python clean_data.py
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

RAW_PATH = "ApexPlanet_DataAnalytics_Dataset.xlsx"
OUTPUT_PATH = "Sales_Dataset_Cleaned.xlsx"

# ---------------------------------------------------------------------------
# STEP 1: Load data
# ---------------------------------------------------------------------------
df = pd.read_excel(RAW_PATH, sheet_name="Sales_Dataset")
print(f"Loaded {len(df)} rows, {len(df.columns)} columns.\n")

# ---------------------------------------------------------------------------
# STEP 2: Data Quality Assessment (profiling)
# ---------------------------------------------------------------------------
print("=" * 60)
print("DATA QUALITY PROFILE (before cleaning)")
print("=" * 60)

print("\nMissing values per column:")
print(df.isnull().sum()[df.isnull().sum() > 0])

print(f"\nFully duplicated rows: {df.duplicated().sum()}")

dup_ids = df["Order_ID"].duplicated(keep=False)
print(f"Rows sharing a non-unique Order_ID: {dup_ids.sum()} "
      f"(unique IDs affected: {df.loc[dup_ids, 'Order_ID'].nunique()})")

print(f"\nOrder_Date dtype before cleaning: {df['Order_Date'].dtype} "
      "(text, not a real date)")

# Outlier detection via IQR on Total_Sales
q1, q3 = df["Total_Sales"].quantile([0.25, 0.75])
iqr = q3 - q1
upper_bound = q3 + 1.5 * iqr
n_outliers = (df["Total_Sales"] > upper_bound).sum()
print(f"\nTotal_Sales outliers (> IQR upper bound of ₹{upper_bound:,.2f}): "
      f"{n_outliers}")

# Sanity check: Total_Sales should equal Quantity * Unit_Price
calc_mismatch = ((df["Quantity"] * df["Unit_Price"] - df["Total_Sales"]).abs() > 1).sum()
print(f"\nRows where Total_Sales != Quantity * Unit_Price: {calc_mismatch}")

# Check categorical fields for casing/whitespace inconsistencies
print("\nCategorical field check (looking for casing/whitespace issues):")
for col in ["Gender", "City", "Product", "Category"]:
    raw_unique = df[col].dropna().unique()
    normalized_unique = set(v.strip().lower() for v in raw_unique)
    flag = "OK" if len(raw_unique) == len(normalized_unique) else "ISSUE FOUND"
    print(f"  {col}: {len(raw_unique)} raw values, "
          f"{len(normalized_unique)} normalized -> {flag}")

# ---------------------------------------------------------------------------
# STEP 3: Data Cleaning & Transformation
# ---------------------------------------------------------------------------
df["Data_Quality_Flag"] = ""

# --- Fix missing Age: impute with median ---
median_age = df["Age"].median()
age_missing_mask = df["Age"].isna()
df.loc[age_missing_mask, "Age"] = median_age
df.loc[age_missing_mask, "Data_Quality_Flag"] += "Age imputed; "

# --- Fix missing City: impute as "Unknown" rather than guessing ---
city_missing_mask = df["City"].isna()
df.loc[city_missing_mask, "City"] = "Unknown"
df.loc[city_missing_mask, "Data_Quality_Flag"] += "City imputed; "

# --- Fix duplicate Order_IDs: keep the first occurrence's ID,
#     re-assign new unique sequential IDs to the rest ---
dup_mask = df["Order_ID"].duplicated(keep="first")
n_dups = dup_mask.sum()
if n_dups > 0:
    max_numeric_id = (
        df["Order_ID"].str.replace("ORD", "", regex=False).astype(int).max()
    )
    new_ids = [f"ORD{max_numeric_id + 1 + i}" for i in range(n_dups)]
    df.loc[dup_mask, "Order_ID"] = new_ids
    df.loc[dup_mask, "Data_Quality_Flag"] += "Order_ID corrected (was duplicate); "

# --- Standardize Order_Date to a real datetime type ---
df["Order_Date"] = pd.to_datetime(df["Order_Date"], errors="coerce")

# --- Feature engineering ---
df["Order_Year"] = df["Order_Date"].dt.year
df["Order_Month"] = df["Order_Date"].dt.month
df["Order_Month_Name"] = df["Order_Date"].dt.strftime("%b")
df["Order_Quarter"] = "Q" + df["Order_Date"].dt.quarter.astype(str)

age_bins = [17, 25, 35, 45, 55, 65]
age_labels = ["18-25", "26-35", "36-45", "46-55", "56-65"]
df["Age_Group"] = pd.cut(df["Age"], bins=age_bins, labels=age_labels)

df["Is_High_Value_Order"] = df["Total_Sales"] > upper_bound

df["Data_Quality_Flag"] = df["Data_Quality_Flag"].str.rstrip("; ")
df.loc[df["Data_Quality_Flag"] == "", "Data_Quality_Flag"] = "Clean"

# ---------------------------------------------------------------------------
# STEP 4: Output the final, analysis-ready dataset
# ---------------------------------------------------------------------------
df.to_excel(OUTPUT_PATH, index=False, sheet_name="Sales_Dataset_Cleaned")

# --- Fix formatting: date display format + column widths + bold header ---
# (Without this, Excel shows "####" for the date column because the default
#  column width is too narrow for the datetime format pandas applies.)
wb = load_workbook(OUTPUT_PATH)
ws = wb["Sales_Dataset_Cleaned"]

date_col_letter = get_column_letter(df.columns.get_loc("Order_Date") + 1)
for row in range(2, ws.max_row + 1):
    ws[f"{date_col_letter}{row}"].number_format = "YYYY-MM-DD"

for col_idx, col_name in enumerate(df.columns, start=1):
    col_letter = get_column_letter(col_idx)
    max_len = max(len(str(col_name)), df[col_name].astype(str).map(len).max())
    ws.column_dimensions[col_letter].width = min(max_len + 3, 30)
    ws[f"{col_letter}1"].font = Font(bold=True)

ws.freeze_panes = "A2"
wb.save(OUTPUT_PATH)

print("\n" + "=" * 60)
print("CLEANING COMPLETE")
print("=" * 60)
print(f"Median age used for imputation: {median_age}")
print(f"Rows with imputed Age: {age_missing_mask.sum()}")
print(f"Rows with imputed City: {city_missing_mask.sum()}")
print(f"Order_IDs corrected: {n_dups}")
print(f"Rows flagged as high-value outliers: {df['Is_High_Value_Order'].sum()}")
print(f"Final dataset shape: {df.shape}")
print(f"Saved to: {OUTPUT_PATH}")
